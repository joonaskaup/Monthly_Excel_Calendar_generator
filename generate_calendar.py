import pandas as pd
import calendar
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
INPUT_FILE = 'Calendar_table.xlsx'  # Update to your input Excel file name
OUTPUT_FILE = 'calendar_output_vertical.xlsx'  # Path for the output Excel file

# Phase to Color Mapping (Added 'Pre-production')
PHASE_COLORS = {
    'Development': 'FFFF00',         # Yellow
    'Pre-pre-production': 'FFA500',  # Orange
    'Pre-production': '83F28F',      # Light Green
    'Shooting': '00C04B',            # Green
    'Post production': '7C4700',     # Brown
    # Add more phases as needed
}

# Read the input Excel file
df = pd.read_excel(INPUT_FILE)

# Display the dataframe to verify correct reading
print("Input DataFrame:")
print(df)

# Display the data types to ensure dates are parsed correctly
print("\nData Types:")
print(df.dtypes)

# Convert 'Start' and 'End' to datetime, then to date objects
df['Start'] = pd.to_datetime(df['Start'], errors='coerce').dt.date
df['End'] = pd.to_datetime(df['End'], errors='coerce').dt.date

# Remove any rows where 'Start' is NaT (invalid or missing dates)
df = df[df['Start'].notna()]

# Handle missing 'End' dates by assuming single-day events
df['End'] = df['End'].fillna(df['Start'])

# Determine the overall date range
min_date = df['Start'].min()
max_date = df['End'].max()

print(f"\nDate Range: {min_date} to {max_date}")

# Create a list of all months in the range
current = datetime(min_date.year, min_date.month, 1)
end = datetime(max_date.year, max_date.month, 1)
months = []
while current <= end:
    months.append((current.year, current.month))
    print(f"Adding month: {current.strftime('%B %Y')}")
    if current.month == 12:
        current = datetime(current.year + 1, 1, 1)
    else:
        current = datetime(current.year, current.month + 1, 1)

print("\nMonths to be included in the calendar:")
for yr, mn in months:
    print(f"{calendar.month_name[mn]} {yr}")

# Initialize the workbook
wb = Workbook()
ws = wb.active
ws.title = "Calendar"

# Define styles
month_header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
weekday_header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Grey
weekend_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Grey for weekends
phase_fill_default = PatternFill(fill_type=None)  # No fill by default

header_font = Font(bold=True, color='000000')  # Black bold font
month_font = Font(bold=True, color='FFFFFF', size=28)  # White bold font, size 28
date_font = Font(bold=True, color='000000', size=12)  # Black bold font for dates
event_font = Font(color='000000', size=10)  # Black font for events

center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Define border style
thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# Set consistent column widths
for col in range(1, 8):  # Columns 1 to 7 (Mon to Sun)
    ws.column_dimensions[get_column_letter(col)].width = 15  # Adjust as needed

# Initialize current_row
current_row = 1

for year, month in months:
    print(f"\nProcessing {calendar.month_name[month]} {year}")
    
    # Write the month and year as header
    month_name = f"{calendar.month_name[month].upper()} {year}"
    month_cell = ws.cell(row=current_row, column=1, value=month_name)
    month_cell.font = month_font
    month_cell.fill = month_header_fill
    month_cell.alignment = center_alignment
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    
    # Set month header row height to 50
    ws.row_dimensions[current_row].height = 50  # Increased row height for month header
    
    # Apply border to the merged month header
    for col in range(1, 8):
        cell = ws.cell(row=current_row, column=col)
        cell.border = thin_border
    
    # Increment row
    current_row += 1
    
    # Write the weekdays headers
    for i, day in enumerate(['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN'], start=1):
        cell = ws.cell(row=current_row, column=i, value=day)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = weekday_header_fill
        cell.border = thin_border
        
        # If the day is Saturday or Sunday, apply weekend fill
        if day in ['SAT', 'SUN']:
            cell.fill = weekend_fill
    
    # Increment row
    current_row += 1
    
    # Generate the calendar for the month
    cal = calendar.Calendar(firstweekday=0)  # Monday as first day
    month_days = cal.monthdatescalendar(year, month)
    
    for week_num, week in enumerate(month_days, start=1):
        print(f"  Processing week {week_num} in {calendar.month_name[month]} {year}")
        
        # Date Number Row
        date_row = current_row
        for day_num, day in enumerate(week, start=1):
            cell = ws.cell(row=date_row, column=day_num)
            if day.month == month:
                cell.value = day.day
                cell.font = date_font
            else:
                cell.value = ""
                cell.fill = phase_fill_default  # No fill for non-current month dates
            cell.alignment = center_alignment
            
            # Determine if the date falls within any event phase
            events_on_day = df[(df['Start'] <= day) & (df['End'] >= day)]
            event_phase = None
            if not events_on_day.empty:
                # Assuming one phase per date, or pick the first one
                event_phase = events_on_day.iloc[0]['Phase']
                print(f"    Date {day} is within event phase: {event_phase}")
            
            # Apply phase color if applicable
            if day.month == month and event_phase and event_phase in PHASE_COLORS:
                cell.fill = PatternFill(start_color=PHASE_COLORS[event_phase], end_color=PHASE_COLORS[event_phase], fill_type='solid')
                print(f"      Applied phase color: {PHASE_COLORS[event_phase]} to date {day}")
            else:
                # Apply weekend grey fill if it's Saturday or Sunday and part of the current month
                if day_num in [6, 7] and day.month == month:
                    cell.fill = weekend_fill
                    print(f"      Applied weekend grey fill to date {day}")
                else:
                    cell.fill = phase_fill_default  # No fill for weekdays without events
            
            # Apply borders only if the cell has a date
            if day.month == month:
                cell.border = thin_border
                ws.row_dimensions[date_row].height = 20  # Set row height for date row (smaller)
            else:
                cell.border = Border()  # No border for non-date cells
                ws.row_dimensions[date_row].height = 20  # Maintain row height
    
        # Increment row
        current_row += 1
        
        # Event Info Row
        event_row = current_row
        for day_num, day in enumerate(week, start=1):
            cell = ws.cell(row=event_row, column=day_num)
            if day.month == month:
                # Find all events for this date
                events = df[(df['Start'] <= day) & (df['End'] >= day)]
                event_titles = [event['Title'] for _, event in events.iterrows()]
                event_phases = [event['Phase'] for _, event in events.iterrows()]
                
                # Omit titles that match their phase (case-insensitive)
                event_texts_filtered = [
                    title for title, phase in zip(event_titles, event_phases)
                    if pd.notna(title) and pd.notna(phase) and title.strip().lower() != phase.strip().lower()
                ]
                cell.value = "\n".join(event_texts_filtered) if event_texts_filtered else ""
                
                # Reset fill to no fill
                cell.fill = phase_fill_default  # Ensures that empty info boxes have no fill
                
                # Apply weekend fill first
                if day_num in [6, 7]:
                    cell.fill = weekend_fill
                    print(f"      Grayed out event info cell for date {day} (Weekend)")
                else:
                    # Apply phase color only if there are filtered events
                    if event_texts_filtered:
                        phase = events.iloc[0]['Phase']
                        if phase in PHASE_COLORS:
                            cell.fill = PatternFill(start_color=PHASE_COLORS[phase], end_color=PHASE_COLORS[phase], fill_type='solid')
                            print(f"      Applied phase color: {PHASE_COLORS[phase]} to event info cell for date {day}")
                
                cell.font = event_font
                cell.alignment = left_alignment
                cell.border = thin_border
                ws.row_dimensions[event_row].height = 40  # Set row height for event row
            else:
                cell.value = ""
                cell.fill = phase_fill_default  # No fill for non-current month dates
                cell.alignment = left_alignment
                cell.border = Border()
                ws.row_dimensions[event_row].height = 40  # Maintain row height
    
        # Increment row after event row
        current_row += 1

# Save the workbook after processing all months
wb.save(OUTPUT_FILE)
print(f"\nVertical calendar with enhanced formatting has been successfully created and saved to {OUTPUT_FILE}")